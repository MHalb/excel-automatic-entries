import re
import openpyxl.styles
from pdfminer.high_level import extract_text
import openpyxl
import warnings
import json
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')



class PDF_CONTROL:

    def __init__(self):
        pass

    def mine_informations(self):
        text =  extract_text("05.pdf")
        lote_list = re.findall(r"\b\d{6}\b", text)
        return lote_list
    



class EXCEL_CONTROL:
    def __init__(self, lote_list, **kwargs):
        self.motorista = "mauro"
        self.exped = "1/5/2024"
        self.truck = "BAS-0431"
        self.sended_location = "CDB"

        self.sended_dep_column = "J"
        self.date_column = "L"
        self.driver_column = "M"
        self.truck_column = "N"
        self.index_culoumn = "O"
        self.lote_column = "F"
        self.code_column = "D"
        
        try:
            self.workbook = openpyxl.load_workbook("PRODUTO ACABADO.xlsm", read_only=False)
        except Exception:
            print("feche todas as planilhas ou remove o arquivo lock para continuar com a execução.")
            exit()
        
        self.planilha = self.workbook.active

        self.metas = {
            "lotes": [int(x) for x in lote_list],
            "coordenadas": {}
        }
        
    def get_metadata_from_excel(self):
        actual_lote = ""
        lock_for_new_search = False
        last_row = ""
        full_palet_quantity = 105



        for row in self.planilha.iter_rows(min_col=4, max_col=10, values_only=False):
            
            if row[2].value in self.metas["lotes"] and row[2].value not in self.metas['coordenadas'] and lock_for_new_search == False:
                actual_lote = row[2].value
                lock_for_new_search = True
                self.metas['coordenadas'][actual_lote] = {
                    'code': row[0].value,
                    'product': row[1].value,
                    'lote': row[2].value,
                    'first_ocurrence': row[2].coordinate,
                    'row_reference_counted': 0,
                    'free_rows': [],
                    'sobras': []
                }


                
            elif row[2].value != actual_lote and lock_for_new_search == True:
                lock_for_new_search = False
                self.metas["coordenadas"][actual_lote]["last_ocurrence"] = last_row[2].coordinate
            
            last_row = row


            if last_row[6].value == "COSMETICO" and lock_for_new_search == True:

                if row[4].value == full_palet_quantity:
                    self.metas["coordenadas"][actual_lote]['free_rows'].append(row[2].coordinate) 
                else:
                    self.metas["coordenadas"][actual_lote]['sobras'].append((row[2].coordinate, row[4].value))

                
                self.metas["coordenadas"][actual_lote]["row_reference_counted"] += 1        


    def recive_crates_quantity(self):
        self.brute_qeue = []
    
        for product in self.metas["coordenadas"].items():
            while 1:
                try:
                    total = int(input("{}, {}, {}".format(
                            product[0], 
                            product[1]['code'], 
                            "insira a quantia de pallets para baixa: "
                        )))
                except ValueError:
                    print("é preciso usar um valor inteiro de 0 até 9000 ou mais, valores como 0 e letas de a-Z são inválidos.")
                    continue


                if total != 0:
                    self.brute_qeue.append((
                        product[0],
                        total)
                    )
                    break
                
                else:
                    print("quantia inválida")

    def upload_informations_to_excel(self):
        for queue in self.brute_qeue:
            content = self.metas["coordenadas"][queue[0]]
            
            for counter in range(int(queue[1])):
                if int(queue[1]) > self.metas["coordenadas"][queue[0]]["row_reference_counted"]:
                    print("valor maior do que a quantia livre, verifique a coluna {} para ver se as colunas e linhas estão corretas.".format(self.metas["coordenadas"][queue[0]]["first_ocurrence"]))
                    break

                actual_row = content['free_rows'][counter]
                
                self.planilha[self.date_column][int(actual_row[1:])-1].value = self.exped
                self.planilha[self.sended_dep_column][int(actual_row[1:])-1].value = self.sended_location
                self.planilha[self.truck_column][int(actual_row[1:])-1].value = self.truck
                self.planilha[self.driver_column][int(actual_row[1:])-1].value = self.motorista
                self.planilha[self.index_culoumn][int(actual_row[1:])-1].value = counter+1

                for location in 'abcdefghijklmno'.upper():
                    self.planilha[f"{location}{int(actual_row[1:])}"].fill = openpyxl.styles.PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    
    def save_excel_changes(self):
        self.workbook.save("PRODUTO ACABADO.xlsm")
    
    def close_workbook(self):
        self.workbook.close()



manager = EXCEL_CONTROL(lote_list=PDF_CONTROL().mine_informations())
manager.get_metadata_from_excel()
manager.recive_crates_quantity()
manager.upload_informations_to_excel()
manager.save_excel_changes()
manager.close_workbook()