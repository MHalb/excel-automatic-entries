import re
import openpyxl.styles
from pdfminer.high_level import extract_text
import openpyxl
import warnings
import json

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

lote_list = []
text =  extract_text("05.pdf")
lote_list = re.findall(r"\b\d{6}\b", text)

workbook = openpyxl.load_workbook("PRODUTO ACABADO.xlsm", read_only=False)
planilha = workbook.active

metas = {
    "lotes": [int(x) for x in lote_list],
    "coordenadas": {}
}



actual_lote = ""
lock_for_new_search = False
lote_first_ocurrence = ""
last_ocurrence = ""
last_row = ""
last_search = False
full_palet_quantity = 105

motorista = "mauro"
exped = "1/5/2024"
truck = "BAS-0431"
sended_location = "CDB"

sended_dep_column = "J"
date_column = "L"
driver_column = "M"
truck_column = "N"
index_culoumn = "O"
lote_column = "F"
code_column = "D"



for row in planilha.iter_rows(min_col=4, max_col=10, values_only=False):
    
    if row[2].value in metas["lotes"] and row[2].value not in metas['coordenadas'] and lock_for_new_search == False:
        actual_lote = row[2].value
        lock_for_new_search = True
        metas['coordenadas'][actual_lote] = {
            'code': row[0].value,
            'product': row[1].value,
            'lote': row[2].value,
            'first_ocurrence': row[2].coordinate,
            'row_reference_counted': 0,
            'free_rows': [],
            'sobras': []
        }


        
    elif row[2].value != actual_lote and lock_for_new_search == True:
        lock_for_new_search = False
        metas["coordenadas"][actual_lote]["last_ocurrence"] = last_row[2].coordinate
    
    last_row = row


    if last_row[6].value == "COSMETICO" and lock_for_new_search == True:

        if row[4].value == full_palet_quantity:
            metas["coordenadas"][actual_lote]['free_rows'].append(row[2].coordinate) 
        else:
            metas["coordenadas"][actual_lote]['sobras'].append((row[2].coordinate, row[4].value))

        
        metas["coordenadas"][actual_lote]["row_reference_counted"] += 1

quantity = []

for product in metas["coordenadas"].items():
    quantity.append((
        product[0], 
        input("{}, {}, {}".format(
            product[0], 
            product[1]['code'], 
            "insira a quantia de pallets para baixa: "
        ))
    ))
    
## na interface fazer um esquema de palet e caixas, caso seja caixas, procurar por pallets não fechados, caso seja palet, procurar por palets fechados



for queue in quantity:
    content = metas["coordenadas"][queue[0]]
    
    for counter in range(int(queue[1])):
        if int(queue[1]) > metas["coordenadas"][queue[0]]["row_reference_counted"]:
            print("valor inválido, verifique a coluna {} para ver se as colunas e linhas estão corretas.".format(metas["coordenadas"][queue[0]]["first_ocurrence"]))
            break

        actual_row = content['free_rows'][counter]
        
        planilha[date_column][int(actual_row[1:])-1].value = exped
        planilha[sended_dep_column][int(actual_row[1:])-1].value = sended_location
        planilha[truck_column][int(actual_row[1:])-1].value = truck
        planilha[driver_column][int(actual_row[1:])-1].value = motorista
        planilha[index_culoumn][int(actual_row[1:])-1].value = counter+1

        for location in 'abcdefghijklmno'.upper():
            planilha[f"{location}{int(actual_row[1:])}"].fill = openpyxl.styles.PatternFill(start_color="548235", end_color="548235", fill_type="solid")
            print(location, actual_row)
        

workbook.save("PRODUTO ACABADO.xlsm")
workbook.close()